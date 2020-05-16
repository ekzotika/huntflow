import requests
import openpyxl
import mimetypes
import os

access_token = input("Введите токен: ")
    #'71e89e8af02206575b3b4ae80bf35b6386fe3085af3d4085cbc7b43505084482'
url_base = "https://dev-100-api.huntflow.ru/"

headers = {
    'User-Agent': 'huntflow/0.1 (eliz.moon5@gmail.com)',
    'Host': 'api.huntflow.ru',
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'Authorization': 'Bearer {0}'.format(access_token)
}

#Тестирование авторизованного пользователя
def test():
    url = '{0}me'.format(url_base)

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        return None

print(test())

#Загрузка резюме
def upload_resume(name, content):
    url_resume = '{0}account/6/upload'.format(url_base)
    mimetypes.init()
    ext_data = os.path.splitext(name)
    if len(ext_data) > 1:
        mime_type = mimetypes.types_map.get(ext_data[len(ext_data) - 1]) or 'application/zip'
    else:
        mime_type = 'application/zip'

    headers = {
        'Authorization': 'Bearer {}'.format(access_token),
        'User-Agent': 'App/1.0 (eeee@eeee.ru)',
        "X-File-Parse": "true",
    }
    r = requests.request(
        'POST',
         url_resume,
         files={'file': (name, content, mime_type)},
         headers=headers,
     )
    return r.json()
upload_resume("glibin_resume.doc", "application/msword")
upload_resume("tanskiy_resume.pdf", "application/pdf")
upload_resume("kornienko_resume.doc", "application/msword")
upload_resume("shorin_resume.pdf", "application/pdf")


#Загрузка кандидата
def loadCandidate(path = input("Введите путь к папке с базой: ")):
    #Считывание Excel
    workbook = openpyxl.load_workbook(r'{0}'.format(path))
    sheet = workbook.worksheets[0]

    position1 = sheet.cell(2, 1).value
    fio1 = sheet.cell(2, 2).value
    salary1 = sheet.cell(2, 3).value
    comment1 = sheet.cell(2, 4).value

    position2 = sheet.cell(3, 1).value
    fio2 = sheet.cell(3, 2).value
    salary2 = sheet.cell(3, 3).value
    comment2 = sheet.cell(3, 4).value

    position3 = sheet.cell(4, 1).value
    fio3 = sheet.cell(4, 2).value
    salary3 = sheet.cell(4, 3).value
    comment3 = sheet.cell(4, 4).value

    position4 = sheet.cell(5, 1).value
    fio4 = sheet.cell(5, 2).value
    salary4 = sheet.cell(5, 3).value
    comment4 = sheet.cell(5, 4).value

    #Загрузка кандидатов в базу
    url = '{0}account/6/applicants'.format(url_base)
    json1 = {
        "last_name": fio1,
        "position": position1,
        "money": salary1,
        "externals": [{"auth_type": "NATIVE", "files": [{"id": 61}]}]
    }

    json2 = {
        "last_name": fio2,
        "position": position2,
        "money": salary2,
        "externals": [{"auth_type": "NATIVE", "files": [{"id": 60}]}]
    }

    json3 = {
        "last_name": fio3,
        "position": position3,
        "money": salary3,
        "externals": [{"auth_type": "NATIVE", "files": [{"id": 59}]}]
    }

    json4 = {
        "last_name": fio4,
        "position": position4,
        "money": salary4,
        "externals": [{"auth_type": "NATIVE", "files": [{"id": 58}]}]
    }
    response1 = requests.post(url, headers=headers, json=json1)
    response2 = requests.post(url, headers=headers, json=json2)
    response3 = requests.post(url, headers=headers, json=json3)
    response4 = requests.post(url, headers=headers, json=json4)


    #Прикрепление кандидата к вакансии
    url_v1 = '{0}account/6/applicants/124/vacancy'.format(url_base)
    url_v2 = '{0}account/6/applicants/125/vacancy'.format(url_base)
    url_v3 = '{0}account/6/applicants/126/vacancy'.format(url_base)
    url_v4 = '{0}account/6/applicants/127/vacancy'.format(url_base)

    json_v1 = {
        "vacancy": 9,
        "status": 43,
        "comment": comment1,
        "files": [
            {
                "id": 61
            }
        ],
    }

    json_v2 = {
        "vacancy": 9,
        "status": 44,
        "comment": comment2,
        "files": [
            {
                "id": 60
            }
        ],
    }

    json_v3 = {
        "vacancy": 2,
        "status": 46,
        "comment": comment3,
        "files": [
            {
                "id": 59
            }
        ],
    }

    json_v4 = {
        "vacancy": 2,
        "status": 50,
        "comment": comment4,
        "files": [
            {
                "id": 58
            }
        ],
    }

    response_v1 = requests.post(url_v1, headers=headers, json=json_v1)
    response_v2 = requests.post(url_v2, headers=headers, json=json_v2)
    response_v3 = requests.post(url_v3, headers=headers, json=json_v3)
    response_v4 = requests.post(url_v4, headers=headers, json=json_v4)


loadCandidate()

